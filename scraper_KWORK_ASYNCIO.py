# pip install aiohttp aiofiles pandas openpyxl python-dotenv

import json
import os
import shutil
from datetime import datetime
from pathlib import Path
import aiohttp
import aiofiles
import asyncio
import pandas as pd
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv


class KworkParser:
    def __init__(self):
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.excel_dir = os.path.join(self.script_dir, 'excel')
        self.session = aiohttp.ClientSession()

    async def fetch(self, url, params=None, cookies=None):
        async with self.session.post(url, data=params, cookies=cookies) as response:
            if response.status == 200:
                return await response.json()
            return None

    async def create_excel_dir(self):
        if os.path.exists(self.excel_dir):
            try:
                shutil.rmtree(self.excel_dir)
                print(f"Removed existing directory: {self.excel_dir}")
            except OSError as e:
                print(f"Error: {e.filename} - {e.strerror}.")

        try:
            os.mkdir(self.excel_dir)
            print(f"Created directory: {self.excel_dir}")
        except OSError as e:
            print(f"Error: {e.filename} - {e.strerror}.")

    async def parse_data(self):
        page_counter = 1
        tasks = []

        async with aiohttp.ClientSession() as session:
            async with session.get('https://kwork.ru/projects') as response:
                response.encoding = 'utf-8'
                cookies = dict(response.cookies)

        for page_counter in range(20):
            url = f'https://kwork.ru/projects?a=1&page={page_counter}'
            params = {'login': 'login', 'pass': 'pass'}
            tasks.append(self.fetch(url, params, cookies))

        responses = await asyncio.gather(*tasks)

        for i, data in enumerate(responses):
            if data:
                data_str = str(data)
                print(data_str.encode("utf-8"))
                print(json.dumps(data))

                df = pd.json_normalize(data)
                df.to_excel(f'kwork_normalize_{i}.xlsx', sheet_name='kwork')
                pd.read_json("data_file.json").to_excel("kwork_output.xlsx")
                print(df)

                df_new = df[["data.wants"]]
                df = pd.DataFrame(df_new, columns=['data.wants'])
                df.to_json('example.json')

                async with aiofiles.open('example.json', 'r') as file:
                    data = json.loads(await file.read())

                book = Workbook()
                sheet = book.active

                sheet['A1'] = 'name'
                sheet['B1'] = 'description'
                sheet['C1'] = 'status'
                sheet['D1'] = 'url'
                sheet['E1'] = 'files'
                sheet['F1'] = 'priceLimit'
                sheet['G1'] = 'possiblePriceLimit'
                sheet['H1'] = 'dateExpire'
                sheet['I1'] = 'dateCreateText'
                sheet['J1'] = 'timeLeft'
                sheet['K1'] = 'userId'

                row = 2
                for j in data['data.wants']['0']:
                    sheet.cell(row=row, column=1).value = str(j['name'])
                    sheet.cell(row=row, column=2).value = str(j['description'])
                    sheet.cell(row=row, column=3).value = str(j['status'])
                    sheet.cell(row=row, column=4).value = 'https://kwork.ru' + '/projects/' + str(j.get('id', '')) + '/view'
                    sheet.cell(row=row, column=5).value = str(j['files'])
                    sheet.cell(row=row, column=6).value = str(j['priceLimit'])
                    sheet.cell(row=row, column=7).value = str(j['possiblePriceLimit'])
                    sheet.cell(row=row, column=8).value = str(j['wantDates'].get('dateExpire', ''))
                    sheet.cell(row=row, column=9).value = str(j['wantDates'].get('dateCreate', ''))
                    sheet.cell(row=row, column=10).value = str(j['timeLeft'])
                    sheet.cell(row=row, column=11).value = str(j.get('wantUserGetProfileUrl', ''))

                    row += 1

                if not os.path.exists('excel'):
                    os.makedirs('excel')

                excel_filename = f'excel/kwork_result_{i}.xlsx'
                book.save(excel_filename)
                book.close()

    async def merge_excel_files(self):
        path = Path("excel")
        min_excel_file_size = 100

        df = pd.concat([pd.read_excel(f)
                        for f in path.glob("*.xlsx")
                        if f.stat().st_size >= min_excel_file_size],
                       ignore_index=True)

        current_date = datetime.now().date()

        df.drop_duplicates(subset=['name'], keep='last', inplace=True)
        df.sort_values(by="name", axis=0, ascending=True, inplace=True)

        output_path = f'excel/kwork_final_{current_date}.xlsx'
        df.to_excel(output_path, index=False)

        wb = load_workbook(output_path)
        ws = wb.active

        ws.freeze_panes = "A2"

        column_widths = {
            'A': 50,
            'B': 55,
            'D': 30,
        }

        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

        wb.save(output_path)

    async def send_to_telegram(self):
        load_dotenv()
        current_date = datetime.now().date()
        BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
        CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")

        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"

        async with aiohttp.ClientSession() as session:
            async with aiofiles.open(f'excel/kwork_final_{current_date}.xlsx', 'rb') as file:
                data = aiohttp.FormData()
                data.add_field('chat_id', CHAT_ID)
                data.add_field('document', await file.read(), filename=f'kwork_final_{current_date}.xlsx')

                async with session.post(url, data=data) as response:
                    if response.status == 200:
                        print("File sent successfully")
                    else:
                        print(f"Failed to send file: {response.status}")

    async def close(self):
        await self.session.close()


async def main():
    kwork_parser = KworkParser()
    await kwork_parser.create_excel_dir()
    await kwork_parser.parse_data()
    await kwork_parser.merge_excel_files()
    await kwork_parser.send_to_telegram()
    await kwork_parser.close()


if __name__ == "__main__":
    asyncio.run(main())
